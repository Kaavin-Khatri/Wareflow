"""Logistics and Financial Document Export Service for PDF and Excel reports.

Generates print-ready PDFs (Pick Lists, Packing Slips, Purchase Orders, Sales Orders, Tax Invoices)
and structured Excel workbooks (Stock Overview, Movement Ledger, AR Aging).
"""

import io
from collections import defaultdict
from datetime import UTC, date, datetime, timedelta
from typing import Any

import openpyxl
from fastapi import HTTPException, status
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import (
    HRFlowable,
    KeepTogether,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.models.retailer import SalesOrder
from app.repositories.interfaces.business_settings_repository import (
    BusinessSettingsRepositoryInterface,
)
from app.repositories.interfaces.delivery_repository import DeliveryRepositoryInterface
from app.repositories.interfaces.invoice_repository import InvoiceRepositoryInterface
from app.repositories.interfaces.product_repository import ProductRepositoryInterface
from app.repositories.interfaces.purchase_order_repository import (
    PurchaseOrderRepositoryInterface,
)
from app.repositories.interfaces.retailer_repository import RetailerRepository
from app.repositories.interfaces.sales_order_repository import SalesOrderRepositoryInterface
from app.repositories.interfaces.supplier_repository import SupplierRepositoryInterface


class ExportService:
    """Document export service generating print-ready PDF reports and structured Excel workbooks."""

    def __init__(
        self,
        sales_order_repo: SalesOrderRepositoryInterface,
        business_settings_repo: BusinessSettingsRepositoryInterface | None = None,
        delivery_repo: DeliveryRepositoryInterface | None = None,
        stock_repo: Any | None = None,
        purchase_order_repo: PurchaseOrderRepositoryInterface | None = None,
        invoice_repo: InvoiceRepositoryInterface | None = None,
        product_repo: ProductRepositoryInterface | None = None,
        retailer_repo: RetailerRepository | None = None,
        supplier_repo: SupplierRepositoryInterface | None = None,
        ar_aging_service: Any | None = None,
    ) -> None:
        self.sales_order_repo = sales_order_repo
        self.business_settings_repo = business_settings_repo
        self.delivery_repo = delivery_repo
        self.stock_repo = stock_repo
        self.purchase_order_repo = purchase_order_repo
        self.invoice_repo = invoice_repo
        self.product_repo = product_repo
        self.retailer_repo = retailer_repo
        self.supplier_repo = supplier_repo
        self.ar_aging_service = ar_aging_service

    def _get_business_profile(self) -> dict[str, str]:
        """Extract business identity and regulatory credentials."""
        settings = self.business_settings_repo.get_settings() if self.business_settings_repo else None
        if settings:
            return {
                "name": getattr(settings, "business_name", None) or "WareFlow Wholesale Logistics Hub",
                "gstin": getattr(settings, "gstin", None) or "27AAAAA0000A1Z5",
                "fssai": getattr(settings, "fssai_license_no", None) or "11521000000123",
                "address": getattr(settings, "address", None) or "Building 4, Bhiwandi Logistics Park, Maharashtra 421302",
                "phone": getattr(settings, "phone", None) or "+91 22 2789 0000",
                "email": getattr(settings, "email", None) or "billing@wareflow.in",
            }
        return {
            "name": "WareFlow Wholesale Logistics Hub",
            "gstin": "27AAAAA0000A1Z5",
            "fssai": "11521000000123",
            "address": "Building 4, Bhiwandi Logistics Park, Maharashtra 421302",
            "phone": "+91 22 2789 0000",
            "email": "billing@wareflow.in",
        }

    def _get_buyer_details(self, order: SalesOrder) -> dict[str, str]:
        """Extract buyer name, contact, phone, and destination address."""
        if getattr(order, "retailer", None):
            return {
                "name": order.retailer.name or "Wholesale Retailer",
                "contact": order.retailer.contact_person or "—",
                "phone": order.retailer.phone or "—",
                "address": order.retailer.address or "Registered Retailer Location",
                "gstin": getattr(order.retailer, "gstin", None) or "—",
                "type": "B2B Retailer",
            }
        if getattr(order, "customer", None):
            return {
                "name": order.customer.name or "Direct Buyer",
                "contact": order.customer.name or "—",
                "phone": order.customer.phone or "—",
                "address": order.customer.address or "Direct Delivery Address",
                "gstin": "Unregistered / Consumer",
                "type": "Direct Customer",
            }
        return {
            "name": getattr(order, "retailer_name", None) or "Direct Buyer",
            "contact": "—",
            "phone": "—",
            "address": "Warehouse Pickup / Delivery Destination",
            "gstin": "—",
            "type": "Buyer",
        }

    def _resolve_uom(self, item: Any) -> str:
        """Safely extract unit of measure string from an item."""
        if getattr(item, "uom", None):
            return (
                getattr(item.uom, "abbreviation", None)
                or getattr(item.uom, "symbol", None)
                or getattr(item.uom, "name", None)
                or "PCS"
            )
        if getattr(item, "product", None):
            return getattr(item.product, "unit", "PCS") or "PCS"
        return "PCS"

    # =========================================================================
    # 1. WAREHOUSE PICK LIST (PDF)
    # =========================================================================
    def generate_pick_list(self, sales_order_id: str) -> bytes:
        """Generate staff-facing Warehouse Pick List PDF (Zero Pricing)."""
        order = self.sales_order_repo.get_by_id(sales_order_id)
        if not order:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Sales order '{sales_order_id}' not found",
            )

        buyer = self._get_buyer_details(order)
        delivery = (
            self.delivery_repo.get_by_sales_order_id(sales_order_id)
            if self.delivery_repo
            else None
        )

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=24,
            rightMargin=24,
            topMargin=24,
            bottomMargin=24,
            pageCompression=0,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "PickTitle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=18,
            textColor=colors.HexColor("#0F172A"),
        )
        subtitle_style = ParagraphStyle(
            "PickSubTitle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#475569"),
        )
        header_cell = ParagraphStyle(
            "HeaderCell",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.white,
            alignment=TA_CENTER,
        )
        normal_text = ParagraphStyle(
            "NormalText",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#0F172A"),
        )
        bold_text = ParagraphStyle(
            "BoldText",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#0F172A"),
        )

        elements = []
        biz = self._get_business_profile()

        # Header Block
        header_table = Table(
            [
                [
                    Paragraph(f"<b>{biz['name']}</b><br/>{biz['address']}", subtitle_style),
                    Paragraph(
                        f"<b>WAREHOUSE PICK LIST</b><br/>"
                        f"<font size=12><b>{order.so_number}</b></font><br/>"
                        f"Order Date: {order.order_date.strftime('%Y-%m-%d %H:%M') if hasattr(order, 'order_date') and order.order_date else datetime.now(UTC).strftime('%Y-%m-%d')}",
                        ParagraphStyle("RightHead", parent=title_style, alignment=TA_RIGHT),
                    ),
                ]
            ],
            colWidths=[300, 245],
        )
        header_table.setStyle(
            TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ])
        )
        elements.append(header_table)
        elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#CBD5E1"), spaceBefore=6, spaceAfter=8))

        # Order & Dispatch Metadata
        vehicle_no = (
            getattr(delivery, "vehicle_no", getattr(delivery, "vehicle_number", "Standard Wholesale Dispatch"))
            if delivery
            else "Standard Wholesale Dispatch"
        )
        meta_table = Table(
            [
                [
                    Paragraph(f"<b>Consignee:</b> {buyer['name']} ({buyer['type']})<br/><b>Destination:</b> {buyer['address']}", normal_text),
                    Paragraph(
                        f"<b>Status:</b> {str(order.status).upper()}<br/>"
                        f"<b>Vehicle:</b> {vehicle_no}",
                        normal_text,
                    ),
                ]
            ],
            colWidths=[330, 215],
        )
        meta_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ])
        )
        elements.append(meta_table)
        elements.append(Spacer(1, 8))

        # Group items by warehouse if present
        items_by_wh = defaultdict(list)
        for it in getattr(order, "items", []):
            wh_name = "Primary Warehouse"
            if getattr(it, "product", None):
                wh = getattr(it.product, "warehouse", None)
                if wh and getattr(wh, "name", None):
                    wh_name = wh.name
                elif getattr(it.product, "warehouse_name", None):
                    wh_name = it.product.warehouse_name
            items_by_wh[wh_name].append(it)

        total_units = 0.0
        global_idx = 1

        for wh_name, wh_items in items_by_wh.items():
            if len(items_by_wh) > 1 or wh_name != "Primary Warehouse":
                elements.append(Paragraph(f"<b>Warehouse Location / Zone: {wh_name}</b>", bold_text))
                elements.append(Spacer(1, 4))

            table_data = [
                [
                    Paragraph("#", header_cell),
                    Paragraph("SKU", header_cell),
                    Paragraph("Product Description", header_cell),
                    Paragraph("Qty to Pick", header_cell),
                    Paragraph("UoM", header_cell),
                    Paragraph("Check", header_cell),
                ]
            ]

            for item in wh_items:
                sku = item.product.sku if getattr(item, "product", None) else "—"
                name = item.product.name if getattr(item, "product", None) else "Product"
                qty = float(item.qty)
                total_units += qty
                uom = self._resolve_uom(item)

                table_data.append([
                    Paragraph(str(global_idx), ParagraphStyle("IdxP", parent=normal_text, alignment=TA_CENTER)),
                    Paragraph(f"<b>{sku}</b>", normal_text),
                    Paragraph(name, normal_text),
                    Paragraph(f"<b>{qty:g}</b>", ParagraphStyle("QtyP", parent=bold_text, alignment=TA_CENTER)),
                    Paragraph(uom, ParagraphStyle("UomP", parent=normal_text, alignment=TA_CENTER)),
                    Paragraph("[ &nbsp; ]", ParagraphStyle("BoxP", parent=bold_text, alignment=TA_CENTER)),
                ])
                global_idx += 1

            items_table = Table(table_data, colWidths=[25, 90, 240, 80, 60, 50])
            items_table.setStyle(
                TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ])
            )
            elements.append(items_table)
            elements.append(Spacer(1, 6))

        # Overall summary row
        summary_table = Table(
            [
                [
                    Paragraph("<b>TOTAL PICK UNITS</b>", bold_text),
                    Paragraph(f"<b>{len(getattr(order, 'items', []))} Line Items</b>", normal_text),
                    Paragraph(f"<b>{total_units:g} Total Units</b>", ParagraphStyle("TotQ", parent=bold_text, alignment=TA_RIGHT)),
                ]
            ],
            colWidths=[150, 195, 200],
        )
        summary_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F1F5F9")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ])
        )
        elements.append(summary_table)
        elements.append(Spacer(1, 10))

        # Sign-off box
        sign_table = Table(
            [
                [
                    Paragraph("<b>PICKED BY:</b> ___________________________<br/>Date & Time: ________________________", normal_text),
                    Paragraph("<b>CHECKED & PACKED BY:</b> ___________________________<br/>Date & Time: ________________________", normal_text),
                ]
            ],
            colWidths=[270, 275],
        )
        sign_table.setStyle(
            TableStyle([
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#94A3B8")),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ])
        )
        elements.append(KeepTogether(sign_table))

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    # =========================================================================
    # 2. CUSTOMER PACKING SLIP (PDF)
    # =========================================================================
    def generate_packing_slip(self, sales_order_id: str) -> bytes:
        """Generate customer-facing Delivery Manifest & Packing Slip PDF (Zero Pricing)."""
        order = self.sales_order_repo.get_by_id(sales_order_id)
        if not order:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Sales order '{sales_order_id}' not found",
            )

        buyer = self._get_buyer_details(order)
        delivery = (
            self.delivery_repo.get_by_sales_order_id(sales_order_id)
            if self.delivery_repo
            else None
        )

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=24,
            rightMargin=24,
            topMargin=24,
            bottomMargin=24,
            pageCompression=0,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "PackTitle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=18,
            textColor=colors.HexColor("#312E81"),
        )
        subtitle_style = ParagraphStyle(
            "PackSubTitle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#475569"),
        )
        header_cell = ParagraphStyle(
            "HeaderCellPack",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.white,
            alignment=TA_CENTER,
        )
        normal_text = ParagraphStyle(
            "NormalTextPack",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#0F172A"),
        )
        bold_text = ParagraphStyle(
            "BoldTextPack",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#0F172A"),
        )

        elements = []
        biz = self._get_business_profile()

        # Header Block
        header_table = Table(
            [
                [
                    Paragraph(
                        f"<b>{biz['name']}</b><br/>"
                        f"GSTIN: {biz['gstin']} | FSSAI: {biz['fssai']}<br/>"
                        f"{biz['address']}<br/>"
                        f"Support: {biz['phone']} | {biz['email']}",
                        subtitle_style,
                    ),
                    Paragraph(
                        f"<b>PACKING SLIP</b><br/>"
                        f"<font size=12><b>{order.so_number}</b></font><br/>"
                        f"Dispatch Date: {datetime.now(UTC).strftime('%Y-%m-%d')}",
                        ParagraphStyle("RightHeadP", parent=title_style, alignment=TA_RIGHT),
                    ),
                ]
            ],
            colWidths=[310, 235],
        )
        elements.append(header_table)
        elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#CBD5E1"), spaceBefore=6, spaceAfter=8))

        # Consignee & Shipping Destination
        vehicle_no = (
            getattr(delivery, "vehicle_no", getattr(delivery, "vehicle_number", "Warehouse Fleet"))
            if delivery
            else "Warehouse Fleet"
        )
        driver_name = getattr(delivery, "driver_name", "Standard Carrier") if delivery else "Standard Carrier"

        meta_table = Table(
            [
                [
                    Paragraph(
                        f"<b>DELIVER TO:</b><br/>"
                        f"<b>{buyer['name']}</b><br/>"
                        f"Contact: {buyer['contact']} ({buyer['phone']})<br/>"
                        f"Address: {buyer['address']}",
                        normal_text,
                    ),
                    Paragraph(
                        f"<b>DISPATCH DETAILS:</b><br/>"
                        f"Order Ref: {order.so_number}<br/>"
                        f"Vehicle: {vehicle_no}<br/>"
                        f"Driver: {driver_name}",
                        normal_text,
                    ),
                ]
            ],
            colWidths=[310, 235],
        )
        meta_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ])
        )
        elements.append(meta_table)
        elements.append(Spacer(1, 8))

        # Line Items
        table_data = [
            [
                Paragraph("#", header_cell),
                Paragraph("SKU", header_cell),
                Paragraph("Item Description", header_cell),
                Paragraph("Qty Shipped", header_cell),
                Paragraph("UoM", header_cell),
                Paragraph("Verified", header_cell),
            ]
        ]

        total_units = 0.0
        for idx, item in enumerate(getattr(order, "items", []), start=1):
            sku = item.product.sku if getattr(item, "product", None) else "—"
            name = item.product.name if getattr(item, "product", None) else "Product"
            qty = float(item.qty)
            total_units += qty
            uom = self._resolve_uom(item)

            table_data.append([
                Paragraph(str(idx), ParagraphStyle("IdxP", parent=normal_text, alignment=TA_CENTER)),
                Paragraph(f"<b>{sku}</b>", normal_text),
                Paragraph(name, normal_text),
                Paragraph(f"<b>{qty:g}</b>", ParagraphStyle("QtyP", parent=bold_text, alignment=TA_CENTER)),
                Paragraph(uom, ParagraphStyle("UomP", parent=normal_text, alignment=TA_CENTER)),
                Paragraph("[ &nbsp;✓&nbsp; ]", ParagraphStyle("VrfP", parent=bold_text, alignment=TA_CENTER)),
            ])

        table_data.append([
            Paragraph("", normal_text),
            Paragraph("<b>TOTAL ITEMS</b>", bold_text),
            Paragraph(f"<b>{len(getattr(order, 'items', []))} Line Items</b>", normal_text),
            Paragraph(f"<b>{total_units:g}</b>", ParagraphStyle("TotQty", parent=bold_text, alignment=TA_CENTER)),
            Paragraph("Units", ParagraphStyle("TotUom", parent=normal_text, alignment=TA_CENTER)),
            Paragraph("", normal_text),
        ])

        items_table = Table(table_data, colWidths=[25, 90, 230, 75, 75, 50])
        items_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4338CA")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F1F5F9")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ])
        )
        elements.append(items_table)
        elements.append(Spacer(1, 10))

        # Consignee Sign-off
        ack_table = Table(
            [
                [
                    Paragraph("<b>CONSIGNEE ACKNOWLEDGMENT:</b>", bold_text),
                    Paragraph("<b>SECURITY GATE PASS:</b>", bold_text),
                ],
                [
                    Paragraph(
                        "Received the above goods in good order and complete condition.<br/><br/>"
                        "Receiver Signature & Stamp: ____________________________<br/>"
                        "Date & Time: __________________________________________",
                        normal_text,
                    ),
                    Paragraph(
                        "Dispatched By: ____________________________________<br/>"
                        "Security Gate Stamp & Signature: __________________",
                        normal_text,
                    ),
                ],
            ],
            colWidths=[310, 235],
        )
        ack_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#94A3B8")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ])
        )
        elements.append(KeepTogether(ack_table))
        elements.append(Spacer(1, 6))

        elements.append(
            Paragraph(
                "<i>Notice: This document is a physical Packing Slip & Delivery Manifest. Official GST Tax Invoice is issued separately.</i>",
                ParagraphStyle("Disclaimer", parent=normal_text, fontSize=7, textColor=colors.HexColor("#64748B"), alignment=TA_CENTER),
            )
        )

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    # =========================================================================
    # 3. PURCHASE ORDER (PDF)
    # =========================================================================
    def generate_purchase_order_pdf(self, purchase_order_id: str) -> bytes:
        """Generate printable Purchase Order document PDF."""
        if not self.purchase_order_repo:
            raise HTTPException(status_code=500, detail="Purchase order repository not configured")

        po = self.purchase_order_repo.get_by_id(purchase_order_id)
        if not po:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Purchase order '{purchase_order_id}' not found",
            )

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=24,
            rightMargin=24,
            topMargin=24,
            bottomMargin=24,
            pageCompression=0,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "POTitle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=18,
            textColor=colors.HexColor("#1E3A8A"),
        )
        subtitle_style = ParagraphStyle(
            "POSubTitle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=11,
            textColor=colors.HexColor("#334155"),
        )
        header_cell = ParagraphStyle(
            "POHeaderCell",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.white,
            alignment=TA_CENTER,
        )
        normal_text = ParagraphStyle(
            "PONormalText",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#0F172A"),
        )
        bold_text = ParagraphStyle(
            "POBoldText",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#0F172A"),
        )
        right_text = ParagraphStyle(
            "PORightText",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            alignment=TA_RIGHT,
            textColor=colors.HexColor("#0F172A"),
        )

        elements = []
        biz = self._get_business_profile()
        supplier = getattr(po, "supplier", None)

        # Header Block
        header_table = Table(
            [
                [
                    Paragraph(
                        f"<b>{biz['name']}</b><br/>"
                        f"GSTIN: {biz['gstin']} | FSSAI: {biz['fssai']}<br/>"
                        f"{biz['address']}<br/>"
                        f"Email: {biz['email']} | Phone: {biz['phone']}",
                        subtitle_style,
                    ),
                    Paragraph(
                        f"<b>PURCHASE ORDER</b><br/>"
                        f"<font size=12><b>{po.po_number}</b></font><br/>"
                        f"PO Date: {po.order_date.strftime('%Y-%m-%d') if hasattr(po, 'order_date') and po.order_date else datetime.now(UTC).strftime('%Y-%m-%d')}<br/>"
                        f"Expected Date: {po.expected_date.strftime('%Y-%m-%d') if getattr(po, 'expected_date', None) else 'Prompt'}",
                        ParagraphStyle("PORightHead", parent=title_style, alignment=TA_RIGHT),
                    ),
                ]
            ],
            colWidths=[310, 235],
        )
        elements.append(header_table)
        elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#CBD5E1"), spaceBefore=6, spaceAfter=8))

        # Vendor / Supplier Info Box
        supp_name = supplier.name if supplier else "Wholesale Supplier / Manufacturer"
        supp_gstin = getattr(supplier, "gstin", None) or "—"
        supp_contact = getattr(supplier, "contact_person", None) or "—"
        supp_phone = getattr(supplier, "phone", None) or "—"
        supp_addr = getattr(supplier, "address", None) or "Registered Vendor Premises"

        vendor_table = Table(
            [
                [
                    Paragraph(
                        f"<b>SUPPLIER / VENDOR:</b><br/>"
                        f"<b>{supp_name}</b><br/>"
                        f"Contact: {supp_contact} | Phone: {supp_phone}<br/>"
                        f"Address: {supp_addr}<br/>"
                        f"GSTIN: {supp_gstin}",
                        normal_text,
                    ),
                    Paragraph(
                        f"<b>DELIVERY & BILLING LOCATION:</b><br/>"
                        f"<b>{biz['name']}</b><br/>"
                        f"Central Receiving Dock, Bhiwandi Hub<br/>"
                        f"Status: <b>{str(po.status).upper()}</b><br/>"
                        f"Payment Terms: Standard 30 Days Credit",
                        normal_text,
                    ),
                ]
            ],
            colWidths=[310, 235],
        )
        vendor_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ])
        )
        elements.append(vendor_table)
        elements.append(Spacer(1, 8))

        # Items Table
        table_data = [
            [
                Paragraph("#", header_cell),
                Paragraph("SKU", header_cell),
                Paragraph("Item Description", header_cell),
                Paragraph("Ordered Qty", header_cell),
                Paragraph("UoM", header_cell),
                Paragraph("Unit Cost", header_cell),
                Paragraph("Total (INR)", header_cell),
            ]
        ]

        total_amount = 0.0
        for idx, item in enumerate(getattr(po, "items", []), start=1):
            sku = item.product.sku if getattr(item, "product", None) else "—"
            name = item.product.name if getattr(item, "product", None) else "Catalog Item"
            qty = float(item.qty_ordered)
            unit_cost = float(item.unit_cost)
            line_tot = qty * unit_cost
            total_amount += line_tot
            uom = self._resolve_uom(item)

            table_data.append([
                Paragraph(str(idx), ParagraphStyle("IdxPO", parent=normal_text, alignment=TA_CENTER)),
                Paragraph(f"<b>{sku}</b>", normal_text),
                Paragraph(name, normal_text),
                Paragraph(f"{qty:g}", ParagraphStyle("QPO", parent=bold_text, alignment=TA_CENTER)),
                Paragraph(uom, ParagraphStyle("UPO", parent=normal_text, alignment=TA_CENTER)),
                Paragraph(f"₹{unit_cost:,.2f}", right_text),
                Paragraph(f"<b>₹{line_tot:,.2f}</b>", ParagraphStyle("TotPO", parent=bold_text, alignment=TA_RIGHT)),
            ])

        # Summary Row
        table_data.append([
            Paragraph("", normal_text),
            Paragraph("<b>TOTAL</b>", bold_text),
            Paragraph(f"<b>{len(getattr(po, 'items', []))} Line Items</b>", normal_text),
            Paragraph("", normal_text),
            Paragraph("", normal_text),
            Paragraph("<b>Grand Total:</b>", ParagraphStyle("TotLbl", parent=bold_text, alignment=TA_RIGHT)),
            Paragraph(f"<b>₹{total_amount:,.2f}</b>", ParagraphStyle("TotVal", parent=bold_text, alignment=TA_RIGHT)),
        ])

        items_table = Table(table_data, colWidths=[25, 85, 195, 65, 45, 65, 65])
        items_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F1F5F9")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ])
        )
        elements.append(items_table)
        elements.append(Spacer(1, 10))

        # Sign-off & Terms Block
        sign_table = Table(
            [
                [
                    Paragraph("<b>PURCHASE ORDER TERMS:</b><br/>1. Goods must match FSSAI quality & batch standards.<br/>2. Dispatches must include delivery challan & original tax invoice.", subtitle_style),
                    Paragraph("<b>AUTHORIZED SIGNATORY:</b><br/><br/>____________________________________<br/>For WareFlow Procurement Division", normal_text),
                ]
            ],
            colWidths=[310, 235],
        )
        sign_table.setStyle(
            TableStyle([
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#94A3B8")),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ])
        )
        elements.append(KeepTogether(sign_table))

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    # =========================================================================
    # 4. SALES ORDER CONFIRMATION (PDF)
    # =========================================================================
    def generate_sales_order_pdf(self, sales_order_id: str) -> bytes:
        """Generate printable Sales Order confirmation PDF with wholesale pricing & tax breakdown."""
        order = self.sales_order_repo.get_by_id(sales_order_id)
        if not order:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Sales order '{sales_order_id}' not found",
            )

        buyer = self._get_buyer_details(order)
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=24,
            rightMargin=24,
            topMargin=24,
            bottomMargin=24,
            pageCompression=0,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "SOTitle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=18,
            textColor=colors.HexColor("#065F46"),
        )
        subtitle_style = ParagraphStyle(
            "SOSubTitle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=11,
            textColor=colors.HexColor("#334155"),
        )
        header_cell = ParagraphStyle(
            "SOHeaderCell",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.white,
            alignment=TA_CENTER,
        )
        normal_text = ParagraphStyle(
            "SONormalText",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#0F172A"),
        )
        bold_text = ParagraphStyle(
            "SOBoldText",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#0F172A"),
        )
        right_text = ParagraphStyle(
            "SORightText",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            alignment=TA_RIGHT,
            textColor=colors.HexColor("#0F172A"),
        )

        elements = []
        biz = self._get_business_profile()

        # Header Block
        header_table = Table(
            [
                [
                    Paragraph(
                        f"<b>{biz['name']}</b><br/>"
                        f"GSTIN: {biz['gstin']} | FSSAI: {biz['fssai']}<br/>"
                        f"{biz['address']}<br/>"
                        f"Phone: {biz['phone']} | Email: {biz['email']}",
                        subtitle_style,
                    ),
                    Paragraph(
                        f"<b>SALES ORDER CONFIRMATION</b><br/>"
                        f"<font size=12><b>{order.so_number}</b></font><br/>"
                        f"Order Date: {order.order_date.strftime('%Y-%m-%d') if hasattr(order, 'order_date') and order.order_date else datetime.now(UTC).strftime('%Y-%m-%d')}<br/>"
                        f"Status: <b>{str(order.status).upper()}</b>",
                        ParagraphStyle("SORightHead", parent=title_style, alignment=TA_RIGHT),
                    ),
                ]
            ],
            colWidths=[310, 235],
        )
        elements.append(header_table)
        elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#CBD5E1"), spaceBefore=6, spaceAfter=8))

        # Buyer Details
        buyer_table = Table(
            [
                [
                    Paragraph(
                        f"<b>BUYER / CONSIGNEE:</b><br/>"
                        f"<b>{buyer['name']}</b> ({buyer['type']})<br/>"
                        f"Contact: {buyer['contact']} | Phone: {buyer['phone']}<br/>"
                        f"GSTIN: {buyer['gstin']}<br/>"
                        f"Delivery Address: {buyer['address']}",
                        normal_text,
                    ),
                    Paragraph(
                        "<b>FULFILLMENT TERMS:</b><br/>"
                        "Dispatch Hub: Bhiwandi Central Facility<br/>"
                        "Payment Terms: Standard Wholesale Credit<br/>"
                        "Delivery Mode: Dedicated Wholesale Freight",
                        normal_text,
                    ),
                ]
            ],
            colWidths=[310, 235],
        )
        buyer_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ])
        )
        elements.append(buyer_table)
        elements.append(Spacer(1, 8))

        # Line items
        table_data = [
            [
                Paragraph("#", header_cell),
                Paragraph("SKU", header_cell),
                Paragraph("Item Description", header_cell),
                Paragraph("Qty", header_cell),
                Paragraph("UoM", header_cell),
                Paragraph("Rate (INR)", header_cell),
                Paragraph("Tax (INR)", header_cell),
                Paragraph("Amount (INR)", header_cell),
            ]
        ]

        subtotal = 0.0
        total_tax = 0.0
        for idx, item in enumerate(getattr(order, "items", []), start=1):
            sku = item.product.sku if getattr(item, "product", None) else "—"
            name = item.product.name if getattr(item, "product", None) else "Catalog Item"
            qty = float(item.qty)
            price = float(item.unit_price)
            tax_amt = float(getattr(item, "tax_amount", 0.0) or 0.0)
            line_tot = float(getattr(item, "total", qty * price + tax_amt))
            subtotal += qty * price
            total_tax += tax_amt
            uom = self._resolve_uom(item)

            table_data.append([
                Paragraph(str(idx), ParagraphStyle("IdxSO", parent=normal_text, alignment=TA_CENTER)),
                Paragraph(f"<b>{sku}</b>", normal_text),
                Paragraph(name, normal_text),
                Paragraph(f"{qty:g}", ParagraphStyle("QSO", parent=bold_text, alignment=TA_CENTER)),
                Paragraph(uom, ParagraphStyle("USO", parent=normal_text, alignment=TA_CENTER)),
                Paragraph(f"₹{price:,.2f}", right_text),
                Paragraph(f"₹{tax_amt:,.2f}", right_text),
                Paragraph(f"<b>₹{line_tot:,.2f}</b>", ParagraphStyle("TotSO", parent=bold_text, alignment=TA_RIGHT)),
            ])

        grand_total = subtotal + total_tax
        table_data.append([
            Paragraph("", normal_text),
            Paragraph("<b>TOTALS</b>", bold_text),
            Paragraph(f"<b>{len(getattr(order, 'items', []))} Line Items</b>", normal_text),
            Paragraph("", normal_text),
            Paragraph("", normal_text),
            Paragraph(f"₹{subtotal:,.2f}", right_text),
            Paragraph(f"₹{total_tax:,.2f}", right_text),
            Paragraph(f"<b>₹{grand_total:,.2f}</b>", ParagraphStyle("GVal", parent=bold_text, alignment=TA_RIGHT)),
        ])

        items_table = Table(table_data, colWidths=[20, 80, 180, 50, 40, 55, 55, 65])
        items_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#065F46")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F1F5F9")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ])
        )
        elements.append(items_table)
        elements.append(Spacer(1, 10))

        # Sign-off & Terms
        sign_table = Table(
            [
                [
                    Paragraph("<b>TERMS & CONDITIONS:</b><br/>1. Subject to Mumbai Jurisdiction.<br/>2. Returns must be requested within 48h of delivery.", subtitle_style),
                    Paragraph("<b>AUTHORIZED SIGNATURE:</b><br/><br/>____________________________________<br/>For WareFlow Commercial Operations", normal_text),
                ]
            ],
            colWidths=[310, 235],
        )
        sign_table.setStyle(
            TableStyle([
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#94A3B8")),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ])
        )
        elements.append(KeepTogether(sign_table))

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    # =========================================================================
    # 5. GST TAX INVOICE (PDF)
    # =========================================================================
    def generate_invoice_pdf(self, invoice_id: str) -> bytes:
        """Generate official print-ready GST Tax Invoice document PDF."""
        if not self.invoice_repo:
            raise HTTPException(status_code=500, detail="Invoice repository not configured")

        invoice = self.invoice_repo.get_by_id(invoice_id)
        if not invoice:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Invoice '{invoice_id}' not found",
            )

        order = getattr(invoice, "sales_order", None)
        buyer = self._get_buyer_details(order) if order else {
            "name": "Wholesale Retailer / Buyer",
            "contact": "—",
            "phone": "—",
            "address": "Registered Retailer Location",
            "gstin": "—",
            "type": "Retailer",
        }

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=24,
            rightMargin=24,
            topMargin=24,
            bottomMargin=24,
            pageCompression=0,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "InvTitle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=18,
            textColor=colors.HexColor("#0F172A"),
        )
        subtitle_style = ParagraphStyle(
            "InvSubTitle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=11,
            textColor=colors.HexColor("#334155"),
        )
        header_cell = ParagraphStyle(
            "InvHeaderCell",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.white,
            alignment=TA_CENTER,
        )
        normal_text = ParagraphStyle(
            "InvNormalText",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#0F172A"),
        )
        bold_text = ParagraphStyle(
            "InvBoldText",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#0F172A"),
        )
        right_text = ParagraphStyle(
            "InvRightText",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            alignment=TA_RIGHT,
            textColor=colors.HexColor("#0F172A"),
        )

        elements = []
        biz = self._get_business_profile()

        # Top Header Banner
        header_table = Table(
            [
                [
                    Paragraph(
                        f"<b>{biz['name']}</b><br/>"
                        f"GSTIN: <b>{biz['gstin']}</b> | State: Maharashtra (27)<br/>"
                        f"FSSAI Lic No: {biz['fssai']}<br/>"
                        f"{biz['address']}<br/>"
                        f"Email: {biz['email']} | Phone: {biz['phone']}",
                        subtitle_style,
                    ),
                    Paragraph(
                        f"<b>TAX INVOICE</b><br/>"
                        f"<font size=8 color='#64748B'>ORIGINAL FOR RECIPIENT</font><br/>"
                        f"<font size=12><b>{invoice.invoice_no}</b></font><br/>"
                        f"Invoice Date: {invoice.invoice_date.strftime('%Y-%m-%d') if hasattr(invoice, 'invoice_date') and invoice.invoice_date else datetime.now(UTC).strftime('%Y-%m-%d')}<br/>"
                        f"Order Ref: {order.so_number if order else 'Direct'}",
                        ParagraphStyle("InvRightHead", parent=title_style, alignment=TA_RIGHT),
                    ),
                ]
            ],
            colWidths=[310, 235],
        )
        elements.append(header_table)
        elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#CBD5E1"), spaceBefore=6, spaceAfter=8))

        # E-Invoice QR / IRN Details if present
        if getattr(invoice, "e_invoice_irn", None):
            e_inv_block = Table(
                [
                    [
                        Paragraph(
                            f"<b>E-INVOICE DETAILS (NIC GST Portal):</b><br/>"
                            f"IRN: <font size=7 fontName='Courier'>{invoice.e_invoice_irn}</font><br/>"
                            f"Ack No: {invoice.e_invoice_ack_no or '—'} | E-Way Bill No: {invoice.e_way_bill_no or '—'}",
                            normal_text,
                        )
                    ]
                ],
                colWidths=[545],
            )
            e_inv_block.setStyle(
                TableStyle([
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EFF6FF")),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#3B82F6")),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ])
            )
            elements.append(e_inv_block)
            elements.append(Spacer(1, 6))

        # Billed To / Consignee Table
        buyer_table = Table(
            [
                [
                    Paragraph(
                        f"<b>BILLED TO & SHIPPED TO:</b><br/>"
                        f"<b>{buyer['name']}</b> ({buyer['type']})<br/>"
                        f"Address: {buyer['address']}<br/>"
                        f"Contact: {buyer['contact']} | Phone: {buyer['phone']}<br/>"
                        f"GSTIN: <b>{buyer['gstin']}</b> | State: Maharashtra (27)",
                        normal_text,
                    ),
                    Paragraph(
                        f"<b>PAYMENT & CREDIT TERMS:</b><br/>"
                        f"Payment Status: <b>{str(invoice.status).upper()}</b><br/>"
                        f"Payment Mode: Bank Transfer / RTGS / UPI<br/>"
                        f"Due Date: {invoice.invoice_date.date() + timedelta(days=30) if hasattr(invoice, 'invoice_date') and invoice.invoice_date else '30 Days Net'}",
                        normal_text,
                    ),
                ]
            ],
            colWidths=[310, 235],
        )
        buyer_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ])
        )
        elements.append(buyer_table)
        elements.append(Spacer(1, 8))

        # Line Items Table
        table_data = [
            [
                Paragraph("#", header_cell),
                Paragraph("Description of Goods", header_cell),
                Paragraph("HSN Code", header_cell),
                Paragraph("Qty", header_cell),
                Paragraph("Rate (INR)", header_cell),
                Paragraph("Taxable Value", header_cell),
                Paragraph("GST", header_cell),
                Paragraph("Total (INR)", header_cell),
            ]
        ]

        subtotal = float(invoice.subtotal or 0.0)
        tax_amount = float(invoice.tax_amount or 0.0)
        total_amount = float(invoice.total_amount or 0.0)

        for idx, item in enumerate(getattr(invoice, "items", []), start=1):
            name = item.product_name or "Wholesale Item"
            hsn = getattr(item, "hsn_code", None) or "1006.30"
            qty = float(item.qty)
            price = float(item.unit_price)
            item_tax = float(getattr(item, "tax_amount", 0.0) or 0.0)
            item_tot = float(getattr(item, "total", qty * price + item_tax))
            tax_rate = float(getattr(item, "tax_rate", 18.0) or 18.0)

            table_data.append([
                Paragraph(str(idx), ParagraphStyle("IdxInv", parent=normal_text, alignment=TA_CENTER)),
                Paragraph(f"<b>{name}</b>", normal_text),
                Paragraph(hsn, ParagraphStyle("HsnInv", parent=normal_text, alignment=TA_CENTER)),
                Paragraph(f"{qty:g}", ParagraphStyle("QInv", parent=bold_text, alignment=TA_CENTER)),
                Paragraph(f"₹{price:,.2f}", right_text),
                Paragraph(f"₹{(qty * price):,.2f}", right_text),
                Paragraph(f"{tax_rate:g}%", ParagraphStyle("TInv", parent=normal_text, alignment=TA_CENTER)),
                Paragraph(f"<b>₹{item_tot:,.2f}</b>", ParagraphStyle("TotInv", parent=bold_text, alignment=TA_RIGHT)),
            ])

        items_table = Table(table_data, colWidths=[20, 180, 55, 45, 60, 65, 45, 75])
        items_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ])
        )
        elements.append(items_table)
        elements.append(Spacer(1, 8))

        # Financial & GST Tax Breakdown Summary
        cgst = tax_amount / 2.0
        sgst = tax_amount / 2.0

        summary_table = Table(
            [
                [
                    Paragraph("<b>BANK DETAILS FOR REMITTANCE:</b><br/>Bank: HDFC Bank Ltd<br/>Account Name: WareFlow Wholesale Logistics<br/>Account No: 50200012345678<br/>IFSC: HDFC0001234<br/>UPI ID: wareflow@hdfcbank", subtitle_style),
                    Paragraph(
                        f"<b>Taxable Subtotal:</b> ₹{subtotal:,.2f}<br/>"
                        f"<b>CGST (9%):</b> ₹{cgst:,.2f}<br/>"
                        f"<b>SGST (9%):</b> ₹{sgst:,.2f}<br/>"
                        f"<b>Total Tax Amount:</b> ₹{tax_amount:,.2f}<br/>"
                        f"<font size=10><b>GRAND TOTAL (INR): ₹{total_amount:,.2f}</b></font>",
                        right_text,
                    ),
                ]
            ],
            colWidths=[300, 245],
        )
        summary_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ])
        )
        elements.append(summary_table)
        elements.append(Spacer(1, 10))

        # Authorized Signatory
        sign_table = Table(
            [
                [
                    Paragraph("<b>DECLARATION:</b><br/>We declare that this invoice shows the actual price of the goods described and that all particulars are true and correct.", subtitle_style),
                    Paragraph(f"<b>For {biz['name']}</b><br/><br/><br/>Authorized Signatory & Official Stamp", ParagraphStyle("SignRight", parent=normal_text, alignment=TA_RIGHT)),
                ]
            ],
            colWidths=[310, 235],
        )
        sign_table.setStyle(
            TableStyle([
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#94A3B8")),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFFFFF")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ])
        )
        elements.append(KeepTogether(sign_table))

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    # =========================================================================
    # 6. STOCK OVERVIEW (EXCEL)
    # =========================================================================
    def generate_stock_overview_excel(self) -> bytes:
        """Generate structured Stock Valuation and Inventory Overview Excel workbook (.xlsx)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock Overview"

        # Theme styles
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        bold_font = Font(name="Calibri", size=11, bold=True)
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )

        headers = [
            "SKU",
            "Product Name",
            "Category",
            "UoM",
            "On-Hand Units",
            "Cost Price (INR)",
            "Wholesale Price (INR)",
            "Total Cost Valuation (INR)",
            "Total Wholesale Valuation (INR)",
            "Reorder Point",
            "Health Status",
        ]

        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        products = self.product_repo.list_all(limit=5000) if self.product_repo else []
        row_idx = 2

        for p in products:
            p_id = str(p.id)
            sku = str(p.sku or "")
            name = str(p.name or "")
            category = getattr(getattr(p, "category", None), "name", "General")
            uom = getattr(p, "unit", "PCS") or "PCS"
            cost_price = float(getattr(p, "cost_price", 0.0) or 0.0)
            wholesale_price = float(getattr(p, "wholesale_price", 0.0) or 0.0)
            reorder_point = float(getattr(p, "reorder_point", 0.0) or 0.0)

            try:
                on_hand = float(self.stock_repo.get_on_hand(p_id)) if self.stock_repo else 0.0
            except Exception:
                on_hand = 0.0

            cost_val = on_hand * cost_price
            ws_val = on_hand * wholesale_price

            if on_hand <= 0:
                status_str = "Critical (Out of Stock)"
            elif reorder_point > 0 and on_hand <= reorder_point:
                status_str = "Low Stock"
            else:
                status_str = "Healthy"

            ws.append([
                sku,
                name,
                category,
                uom,
                on_hand,
                cost_price,
                wholesale_price,
                cost_val,
                ws_val,
                reorder_point,
                status_str,
            ])

            for col_num in range(1, len(headers) + 1):
                c = ws.cell(row=row_idx, column=col_num)
                c.border = thin_border
                if col_num in (5, 10):  # Qty
                    c.number_format = "#,##0"
                elif col_num in (6, 7, 8, 9):  # Currency
                    c.number_format = "₹#,##0.00"

            row_idx += 1

        # Summary Total Row
        if row_idx > 2:
            ws.append([
                "TOTAL",
                f"{len(products)} Products",
                "",
                "",
                f"=SUM(E2:E{row_idx-1})",
                "",
                "",
                f"=SUM(H2:H{row_idx-1})",
                f"=SUM(I2:I{row_idx-1})",
                "",
                "",
            ])
            for col_num in range(1, len(headers) + 1):
                c = ws.cell(row=row_idx, column=col_num)
                c.font = bold_font
                c.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
                if col_num == 5:
                    c.number_format = "#,##0"
                elif col_num in (8, 9):
                    c.number_format = "₹#,##0.00"

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    # =========================================================================
    # 7. STOCK MOVEMENT LEDGER (EXCEL)
    # =========================================================================
    def generate_stock_movements_excel(self, limit: int = 2000) -> bytes:
        """Generate structured Stock Movement Ledger Excel workbook (.xlsx)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Movement Ledger"

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )

        headers = [
            "Movement ID",
            "Date & Time (UTC)",
            "Movement Type",
            "Product SKU",
            "Product Name",
            "Quantity",
            "Reference Type",
            "Reference ID",
        ]

        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        movements = []
        if self.stock_repo and hasattr(self.stock_repo, "list_movements"):
            try:
                movements = self.stock_repo.list_movements(limit=limit)
            except Exception:
                movements = []

        row_idx = 2
        for m in movements:
            m_id = str(getattr(m, "id", ""))
            created_at = getattr(m, "created_at", datetime.now(UTC))
            dt_str = created_at.strftime("%Y-%m-%d %H:%M") if hasattr(created_at, "strftime") else str(created_at)[:16]
            m_type = str(getattr(m, "type", getattr(m, "movement_type", "MOVEMENT"))).upper()
            sku = getattr(getattr(m, "product", None), "sku", "—")
            p_name = getattr(getattr(m, "product", None), "name", "Product")
            qty = float(getattr(m, "quantity", 0.0) or 0.0)
            ref_type = str(getattr(m, "reference_type", "—") or "—")
            ref_id = str(getattr(m, "reference_id", "—") or "—")

            ws.append([
                m_id,
                dt_str,
                m_type,
                sku,
                p_name,
                qty,
                ref_type,
                ref_id,
            ])

            for col_num in range(1, len(headers) + 1):
                c = ws.cell(row=row_idx, column=col_num)
                c.border = thin_border
                if col_num == 6:
                    c.number_format = "#,##0.00"

            row_idx += 1

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    # =========================================================================
    # 8. ACCOUNTS RECEIVABLE AGING (EXCEL)
    # =========================================================================
    def generate_ar_aging_excel(
        self,
        include_zero_balance: bool = True,
        as_of: date | None = None,
    ) -> bytes:
        """Generate structured Accounts Receivable Aging Report Excel workbook (.xlsx)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AR Aging Report"

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        bold_font = Font(name="Calibri", size=11, bold=True)
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )

        headers = [
            "Retailer ID",
            "Retailer Name",
            "Contact Person",
            "Phone",
            "Credit Limit (INR)",
            "Current / Within Terms (INR)",
            "1-30 Days Overdue (INR)",
            "31-60 Days Overdue (INR)",
            "61-90 Days Overdue (INR)",
            "90+ Days Critical Overdue (INR)",
            "Total Overdue (INR)",
            "Total Outstanding (INR)",
            "Oldest Invoice Date",
            "Open Invoices Count",
        ]

        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        report_data = None
        if self.ar_aging_service:
            report_data = self.ar_aging_service.get_ar_aging_report(
                include_zero_balance=include_zero_balance, as_of=as_of
            )

        retailers = report_data.retailers if report_data else []
        row_idx = 2

        for r in retailers:
            ws.append([
                r.retailer_id,
                r.retailer_name,
                r.contact_person or "",
                r.phone or "",
                r.credit_limit,
                r.current,
                r.bucket_1_30,
                r.bucket_31_60,
                r.bucket_61_90,
                r.bucket_90_plus,
                r.total_overdue,
                r.total_outstanding,
                r.oldest_invoice_date or "N/A",
                r.invoice_count,
            ])

            for col_num in range(1, len(headers) + 1):
                c = ws.cell(row=row_idx, column=col_num)
                c.border = thin_border
                if col_num in (5, 6, 7, 8, 9, 10, 11, 12):
                    c.number_format = "₹#,##0.00"
                elif col_num == 14:
                    c.number_format = "#,##0"

            row_idx += 1

        # Summary Total Row
        if row_idx > 2:
            ws.append([
                "PORTFOLIO TOTAL",
                f"{len(retailers)} Retailers",
                "",
                "",
                f"=SUM(E2:E{row_idx-1})",
                f"=SUM(F2:F{row_idx-1})",
                f"=SUM(G2:G{row_idx-1})",
                f"=SUM(H2:H{row_idx-1})",
                f"=SUM(I2:I{row_idx-1})",
                f"=SUM(J2:J{row_idx-1})",
                f"=SUM(K2:K{row_idx-1})",
                f"=SUM(L2:L{row_idx-1})",
                "",
                f"=SUM(N2:N{row_idx-1})",
            ])
            for col_num in range(1, len(headers) + 1):
                c = ws.cell(row=row_idx, column=col_num)
                c.font = bold_font
                c.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
                if col_num in (5, 6, 7, 8, 9, 10, 11, 12):
                    c.number_format = "₹#,##0.00"
                elif col_num == 14:
                    c.number_format = "#,##0"

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
