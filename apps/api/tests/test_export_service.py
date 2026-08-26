"""Unit and Integration Tests for Step 15.3: Excel & PDF Document Export."""

import io
from datetime import UTC, date, datetime
from decimal import Decimal
from unittest.mock import MagicMock

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.core.security import CurrentUser, get_current_user, require_permission
from app.main import app
from app.models.billing import Invoice, InvoiceItem, InvoiceStatusEnum
from app.models.catalog import Category, Product
from app.models.inventory import StockMovement, StockMovementTypeEnum
from app.models.retailer import Retailer, SalesOrder, SalesOrderItem, SOStatusEnum
from app.models.supplier import POStatusEnum, PurchaseOrder, PurchaseOrderItem, Supplier
from app.models.uom import UnitOfMeasure
from app.services.export_service import ExportService


@pytest.fixture
def mock_user():
    return CurrentUser(
        id="usr-admin-1",
        email="owner@wareflow.in",
        role="admin",
        permissions=["invoices:view", "orders:view", "inventory:view", "*"],
    )


@pytest.fixture
def mock_export_deps():
    so_repo = MagicMock()
    biz_repo = MagicMock()
    delivery_repo = MagicMock()
    stock_repo = MagicMock()
    po_repo = MagicMock()
    invoice_repo = MagicMock()
    product_repo = MagicMock()
    retailer_repo = MagicMock()
    supplier_repo = MagicMock()
    ar_aging_service = MagicMock()

    service = ExportService(
        sales_order_repo=so_repo,
        business_settings_repo=biz_repo,
        delivery_repo=delivery_repo,
        stock_repo=stock_repo,
        purchase_order_repo=po_repo,
        invoice_repo=invoice_repo,
        product_repo=product_repo,
        retailer_repo=retailer_repo,
        supplier_repo=supplier_repo,
        ar_aging_service=ar_aging_service,
    )
    return {
        "service": service,
        "so_repo": so_repo,
        "biz_repo": biz_repo,
        "delivery_repo": delivery_repo,
        "stock_repo": stock_repo,
        "po_repo": po_repo,
        "invoice_repo": invoice_repo,
        "product_repo": product_repo,
        "retailer_repo": retailer_repo,
        "supplier_repo": supplier_repo,
        "ar_aging_service": ar_aging_service,
    }


def test_purchase_order_pdf_export(mock_export_deps):
    service = mock_export_deps["service"]
    po_repo = mock_export_deps["po_repo"]

    supplier = Supplier(
        id="supp-1",
        name="Organic Agro Mills Ltd",
        contact_person="Ravi Verma",
        phone="+919811223344",
        email="orders@organicagro.in",
        address="MIDC Industrial Area, Pune 411018",
        gstin="27AABCO1234F1Z5",
    )
    prod = Product(id="prod-1", sku="BAS-1KG", name="Royal Basmati Rice 1kg", unit="BAG")
    uom = UnitOfMeasure(id="uom-1", name="Bag", abbreviation="BAG")
    po_item = PurchaseOrderItem(
        id="poi-1",
        po_id="po-101",
        product_id="prod-1",
        qty_ordered=Decimal("100"),
        unit_cost=Decimal("150.00"),
        product=prod,
        uom=uom,
    )
    po = PurchaseOrder(
        id="po-101",
        po_number="PO-2026-0001",
        supplier_id="supp-1",
        status=POStatusEnum.ORDERED,
        order_date=datetime(2026, 8, 20, 10, 0, tzinfo=UTC),
        expected_date=date(2026, 8, 25),
        total_amount=Decimal("15000.00"),
        supplier=supplier,
        items=[po_item],
    )
    po_repo.get_by_id.return_value = po

    pdf_bytes = service.generate_purchase_order_pdf("po-101")
    assert isinstance(pdf_bytes, bytes)
    assert len(pdf_bytes) > 500
    assert pdf_bytes.startswith(b"%PDF")


def test_sales_order_pdf_export(mock_export_deps):
    service = mock_export_deps["service"]
    so_repo = mock_export_deps["so_repo"]

    retailer = Retailer(
        id="ret-1",
        name="Vashi Wholesale Supermart",
        contact_person="Kiran Shah",
        phone="+919820011223",
        address="Sector 19, APMC Vashi, Navi Mumbai 400703",
        gstin="27AAACV9876Q1Z9",
    )
    prod = Product(id="prod-1", sku="BAS-1KG", name="Royal Basmati Rice 1kg", unit="BAG")
    uom = UnitOfMeasure(id="uom-1", name="Bag", abbreviation="BAG")
    so_item = SalesOrderItem(
        id="soi-1",
        so_id="so-101",
        product_id="prod-1",
        qty=Decimal("50"),
        unit_price=Decimal("200.00"),
        product=prod,
        uom=uom,
    )
    so = SalesOrder(
        id="so-101",
        so_number="SO-2026-0001",
        retailer_id="ret-1",
        status=SOStatusEnum.CONFIRMED,
        order_date=datetime(2026, 8, 20, 11, 0, tzinfo=UTC),
        total_amount=Decimal("10500.00"),
        retailer=retailer,
        items=[so_item],
    )
    so_repo.get_by_id.return_value = so

    pdf_bytes = service.generate_sales_order_pdf("so-101")
    assert isinstance(pdf_bytes, bytes)
    assert len(pdf_bytes) > 500
    assert pdf_bytes.startswith(b"%PDF")


def test_invoice_pdf_export_with_gst_breakdown(mock_export_deps):
    service = mock_export_deps["service"]
    inv_repo = mock_export_deps["invoice_repo"]

    retailer = Retailer(
        id="ret-1",
        name="Vashi Wholesale Supermart",
        contact_person="Kiran Shah",
        phone="+919820011223",
        address="Sector 19, APMC Vashi, Navi Mumbai 400703",
        gstin="27AAACV9876Q1Z9",
    )
    so = SalesOrder(
        id="so-101",
        so_number="SO-2026-0001",
        retailer_id="ret-1",
        retailer=retailer,
    )
    inv_item = InvoiceItem(
        id="inv-item-1",
        invoice_id="inv-101",
        product_name="Royal Basmati Rice 1kg",
        hsn_code="1006.30",
        qty=Decimal("50"),
        unit_price=Decimal("200.00"),
        tax_rate=Decimal("5.0"),
        tax_amount=Decimal("500.00"),
        total=Decimal("10500.00"),
    )
    invoice = Invoice(
        id="inv-101",
        invoice_no="INV-2026-0001",
        sales_order_id="so-101",
        status=InvoiceStatusEnum.UNPAID,
        invoice_date=datetime(2026, 8, 20, 12, 0, tzinfo=UTC),
        subtotal=Decimal("10000.00"),
        tax_amount=Decimal("500.00"),
        total_amount=Decimal("10500.00"),
        sales_order=so,
        items=[inv_item],
        e_invoice_irn="9b8374a5892cbe34120194827103487192837419283741928374192837419283",
        e_invoice_ack_no="122610928374",
        e_way_bill_no="241098273615",
    )
    inv_repo.get_by_id.return_value = invoice

    pdf_bytes = service.generate_invoice_pdf("inv-101")
    assert isinstance(pdf_bytes, bytes)
    assert len(pdf_bytes) > 500
    assert pdf_bytes.startswith(b"%PDF")


def test_stock_overview_excel_export(mock_export_deps):
    service = mock_export_deps["service"]
    prod_repo = mock_export_deps["product_repo"]
    stock_repo = mock_export_deps["stock_repo"]

    cat = Category(id="cat-1", name="Grains & Cereals")
    p1 = Product(
        id="prod-1",
        sku="BAS-1KG",
        name="Royal Basmati Rice 1kg",
        category=cat,
        unit="BAG",
        cost_price=Decimal("120.00"),
        wholesale_price=Decimal("150.00"),
        reorder_point=Decimal("20.0"),
    )
    prod_repo.list_all.return_value = [p1]
    stock_repo.get_on_hand.return_value = Decimal("80.0")

    xlsx_bytes = service.generate_stock_overview_excel()
    assert isinstance(xlsx_bytes, bytes)
    assert len(xlsx_bytes) > 100

    # Parse openpyxl workbook
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Stock Overview"]
    assert ws.cell(row=1, column=1).value == "SKU"
    assert ws.cell(row=1, column=2).value == "Product Name"
    assert ws.cell(row=2, column=1).value == "BAS-1KG"
    assert ws.cell(row=2, column=2).value == "Royal Basmati Rice 1kg"
    assert ws.cell(row=2, column=5).value == 80.0


def test_stock_movements_excel_export(mock_export_deps):
    service = mock_export_deps["service"]
    stock_repo = mock_export_deps["stock_repo"]

    prod = Product(id="prod-1", sku="BAS-1KG", name="Royal Basmati Rice 1kg")
    mov = StockMovement(
        id="mov-1",
        type=StockMovementTypeEnum.IN,
        product_id="prod-1",
        quantity=Decimal("100"),
        reference_type="PO",
        reference_id="po-101",
        product=prod,
        created_at=datetime(2026, 8, 20, 10, 0, tzinfo=UTC),
    )
    stock_repo.list_movements.return_value = [mov]

    xlsx_bytes = service.generate_stock_movements_excel()
    assert isinstance(xlsx_bytes, bytes)
    assert len(xlsx_bytes) > 100

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Movement Ledger"]
    assert ws.cell(row=1, column=1).value == "Movement ID"
    assert ws.cell(row=2, column=1).value == "mov-1"
    assert ws.cell(row=2, column=3).value == "IN"
    assert ws.cell(row=2, column=4).value == "BAS-1KG"


def test_ar_aging_excel_export(mock_export_deps):
    service = mock_export_deps["service"]
    ar_service = mock_export_deps["ar_aging_service"]

    from app.schemas.analytics import ARAgingBucketItem, ARAgingReportResponse, ARAgingSummary

    item = ARAgingBucketItem(
        retailer_id="ret-1",
        retailer_name="Vashi Wholesale Supermart",
        contact_person="Kiran Shah",
        phone="+919820011223",
        credit_limit=500000.0,
        credit_balance=150000.0,
        current=50000.0,
        bucket_1_30=40000.0,
        bucket_31_60=60000.0,
        bucket_61_90=0.0,
        bucket_90_plus=0.0,
        total_overdue=100000.0,
        total_outstanding=150000.0,
        oldest_invoice_date="2026-07-01",
        invoice_count=2,
    )
    summary = ARAgingSummary(
        total_current=50000.0,
        total_bucket_1_30=40000.0,
        total_bucket_31_60=60000.0,
        total_bucket_61_90=0.0,
        total_bucket_90_plus=0.0,
        total_overdue=100000.0,
        total_outstanding=150000.0,
        total_retailers=1,
        overdue_retailers_count=1,
    )
    mock_report = ARAgingReportResponse(
        as_of_date="2026-08-24",
        summary=summary,
        retailers=[item],
        generated_at="2026-08-24T12:00:00Z",
    )
    ar_service.get_ar_aging_report.return_value = mock_report

    xlsx_bytes = service.generate_ar_aging_excel()
    assert isinstance(xlsx_bytes, bytes)
    assert len(xlsx_bytes) > 100

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["AR Aging Report"]
    assert ws.cell(row=1, column=1).value == "Retailer ID"
    assert ws.cell(row=2, column=1).value == "ret-1"
    assert ws.cell(row=2, column=2).value == "Vashi Wholesale Supermart"
    assert ws.cell(row=2, column=7).value == 40000.0


def test_api_export_endpoints(mock_user, monkeypatch):
    """Test all 6 FastAPI document export routes with TestClient."""
    from app.core.di import get_export_service

    mock_service = MagicMock()
    mock_service.generate_purchase_order_pdf.return_value = b"%PDF-1.4 PO content"
    mock_service.generate_sales_order_pdf.return_value = b"%PDF-1.4 SO content"
    mock_service.generate_invoice_pdf.return_value = b"%PDF-1.4 Invoice content"
    mock_service.generate_stock_overview_excel.return_value = b"PK\x03\x04 Excel overview"
    mock_service.generate_stock_movements_excel.return_value = b"PK\x03\x04 Excel movements"
    mock_service.generate_ar_aging_excel.return_value = b"PK\x03\x04 Excel ar aging"

    app.dependency_overrides[get_current_user] = lambda: mock_user
    app.dependency_overrides[require_permission("invoices:view")] = lambda: mock_user
    app.dependency_overrides[get_export_service] = lambda: mock_service

    client = TestClient(app)

    # 1. PO PDF
    r = client.get("/purchase-orders/po-101/pdf")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/pdf"
    assert r.content == b"%PDF-1.4 PO content"

    # 2. SO PDF
    r = client.get("/sales-orders/so-101/pdf")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/pdf"
    assert r.content == b"%PDF-1.4 SO content"

    # 3. Invoice PDF
    r = client.get("/invoices/inv-101/pdf")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/pdf"
    assert r.content == b"%PDF-1.4 Invoice content"

    # 4. Stock Overview Excel
    r = client.get("/stock/overview.xlsx")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert r.content == b"PK\x03\x04 Excel overview"

    # 5. Stock Movements Excel
    r = client.get("/stock/movements.xlsx")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert r.content == b"PK\x03\x04 Excel movements"

    # 6. AR Aging Excel
    r = client.get("/analytics/ar-aging.xlsx")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert r.content == b"PK\x03\x04 Excel ar aging"

    app.dependency_overrides.clear()
